#!/usr/bin/env python3
"""Export reviewer-friendly Excel with merged model predictions.

This script reads:
- original sentence-level Excel workbook
- gold labels CSV (prepared by prepare_excel_prompting_inputs.py)
- model review JSONL (from prompting outputs)

It writes a new workbook with additional columns:
- predicted_label
- prediction_confidence
- prediction_reason
- eval_status (TP/FP/FN/TN)
- distance_to_nearest_gold
- normalized_label_min3
- normalized_label_min5
- normalized_status_min5
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any

import pandas as pd


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm_text(value: Any) -> str:
    return re.sub(r"\s+", " ", clean(value))


def as_label(value: Any) -> str:
    token = clean(value).upper()
    if token in {"BORDER", "TRUE", "YES", "1"}:
        return "BORDER"
    return "NOBORDER"


def coerce_int(value: Any, fallback: int) -> int:
    try:
        return int(float(clean(value)))
    except Exception:
        return fallback


def infer_sentence_id_column(columns: list[str]) -> str:
    lowered = {c.lower(): c for c in columns}
    for preferred in ("satz_id", "sent_nr", "sentence_nr", "sent_id", "sentence_index"):
        if preferred in lowered:
            return lowered[preferred]
    for col in columns:
        lc = col.lower()
        if ("satz" in lc or "sent" in lc) and ("id" in lc or "index" in lc):
            return col
        if lc.endswith("_nr") or lc == "nr":
            return col
    raise ValueError(f"Could not infer sentence index column from: {columns}")


def infer_text_column(columns: list[str]) -> str:
    lowered = {c.lower(): c for c in columns}
    for preferred in ("sent_text", "sentence_text", "text", "sentence_text_full"):
        if preferred in lowered:
            return lowered[preferred]
    for col in columns:
        lc = col.lower()
        if "text" in lc and "rating" not in lc:
            return col
    raise ValueError(f"Could not infer sentence text column from: {columns}")


def parse_confidence(raw_model_response: Any) -> float | None:
    if not isinstance(raw_model_response, str):
        return None
    try:
        payload = json.loads(raw_model_response)
    except Exception:
        return None
    value = payload.get("confidence")
    if value is None:
        return None
    try:
        return float(value)
    except Exception:
        return None


def parse_reason(raw_model_response: Any) -> str:
    if not isinstance(raw_model_response, str):
        return ""
    try:
        payload = json.loads(raw_model_response)
    except Exception:
        return ""
    return clean(payload.get("reason"))


def read_gold_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def read_review_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def apply_min_scene_len(pred_labels: list[str], min_gap: int) -> list[str]:
    out = pred_labels[:]
    last_kept_border = -10**9
    for i, label in enumerate(out):
        if label != "BORDER":
            continue
        if i - last_kept_border < min_gap:
            out[i] = "NOBORDER"
        else:
            last_kept_border = i
    return out


def eval_status(pred: str, gold: str) -> str:
    if pred == "BORDER" and gold == "BORDER":
        return "TP"
    if pred == "BORDER" and gold == "NOBORDER":
        return "FP"
    if pred == "NOBORDER" and gold == "BORDER":
        return "FN"
    return "TN"


def nearest_distance(value: int, indices: list[int]) -> int | None:
    if not indices:
        return None
    return min(abs(value - idx) for idx in indices)


def build_aligned_rows(
    gold_rows: list[dict[str, Any]],
    review_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[int, dict[str, Any]], dict[str, dict[str, Any]]]:
    pred_by_idx: dict[int, dict[str, Any]] = {}
    pred_by_text: dict[str, dict[str, Any]] = {}
    for row in review_rows:
        idx_raw = clean(row.get("sentence_index"))
        if idx_raw:
            try:
                pred_by_idx[int(float(idx_raw))] = row
            except Exception:
                pass
        text_key = norm_text(row.get("sentence_text_full"))
        if text_key and text_key not in pred_by_text:
            pred_by_text[text_key] = row

    aligned: list[dict[str, Any]] = []
    for pos, row in enumerate(gold_rows):
        idx = coerce_int(row.get("sentence_index"), fallback=pos)
        text = clean(row.get("sentence_text_full"))
        gold_label = as_label(row.get("ground_truth_label"))

        pred_row = pred_by_idx.get(idx)
        matched_by = "index"
        if pred_row is None and text:
            pred_row = pred_by_text.get(norm_text(text))
            matched_by = "text"

        if pred_row is None:
            pred_label = "NOBORDER"
            confidence = None
            reason = ""
            matched_by = "default_noborder"
        else:
            pred_label = as_label(pred_row.get("prediction_label"))
            confidence = parse_confidence(pred_row.get("raw_model_response"))
            reason = parse_reason(pred_row.get("raw_model_response"))

        aligned.append(
            {
                "pos": pos,
                "sentence_index": idx,
                "sentence_text_full": text,
                "ground_truth_label": gold_label,
                "predicted_label": pred_label,
                "prediction_confidence": confidence,
                "prediction_reason": reason,
                "match_source": matched_by,
            }
        )

    pred_seq = [row["predicted_label"] for row in aligned]
    min3_seq = apply_min_scene_len(pred_seq, min_gap=3)
    min5_seq = apply_min_scene_len(pred_seq, min_gap=5)
    gold_border_pos = [row["pos"] for row in aligned if row["ground_truth_label"] == "BORDER"]

    by_idx: dict[int, dict[str, Any]] = {}
    by_text: dict[str, dict[str, Any]] = {}
    for row, min3_label, min5_label in zip(aligned, min3_seq, min5_seq):
        dist = None
        if row["predicted_label"] == "BORDER":
            dist = nearest_distance(row["pos"], gold_border_pos)
        enriched = {
            **row,
            "eval_status": eval_status(row["predicted_label"], row["ground_truth_label"]),
            "distance_to_nearest_gold": dist,
            "normalized_label_min3": min3_label,
            "normalized_label_min5": min5_label,
            "normalized_status_min5": eval_status(min5_label, row["ground_truth_label"]),
        }
        by_idx[enriched["sentence_index"]] = enriched
        text_key = norm_text(enriched["sentence_text_full"])
        if text_key and text_key not in by_text:
            by_text[text_key] = enriched

    return aligned, by_idx, by_text


def add_conditional_colors(path: Path, status_columns: list[str]) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except Exception:
        return

    wb = load_workbook(path)
    ws = wb.active

    header_map: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        header_map[clean(cell.value)] = col_idx

    fill_map = {
        "TP": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "FP": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "FN": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "TN": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
    }

    for status_col in status_columns:
        col_idx = header_map.get(status_col)
        if not col_idx:
            continue
        for row_idx in range(2, ws.max_row + 1):
            value = clean(ws.cell(row=row_idx, column=col_idx).value).upper()
            fill = fill_map.get(value)
            if fill is not None:
                ws.cell(row=row_idx, column=col_idx).fill = fill

    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel_file", type=Path, required=True, help="Original source workbook path.")
    parser.add_argument("--gold_csv", type=Path, required=True, help="Gold labels CSV path.")
    parser.add_argument("--review_jsonl", type=Path, required=True, help="Model review JSONL path.")
    parser.add_argument("--sheet_name", type=str, default="Sheet1", help="Sheet to process.")
    parser.add_argument("--out_xlsx", type=Path, required=True, help="Output workbook path.")
    parser.add_argument(
        "--no_colors",
        action="store_true",
        help="Disable conditional status colors in output workbook.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    df = pd.read_excel(args.excel_file, sheet_name=args.sheet_name)
    df.columns = [str(col).strip() for col in df.columns]
    columns = [str(col) for col in df.columns]
    sentence_id_col = infer_sentence_id_column(columns)
    text_col = infer_text_column(columns)

    gold_rows = read_gold_csv(args.gold_csv)
    review_rows = read_review_jsonl(args.review_jsonl)
    _, aligned_by_idx, aligned_by_text = build_aligned_rows(gold_rows, review_rows)

    enriched_rows: list[dict[str, Any]] = []
    for pos, row in df.iterrows():
        row_dict = row.to_dict()
        sentence_index = coerce_int(row_dict.get(sentence_id_col), fallback=pos)
        sentence_text = clean(row_dict.get(text_col))

        aligned = aligned_by_idx.get(sentence_index)
        if aligned is None and sentence_text:
            aligned = aligned_by_text.get(norm_text(sentence_text))

        if aligned is None:
            aligned = {
                "ground_truth_label": "",
                "predicted_label": "",
                "prediction_confidence": None,
                "prediction_reason": "",
                "eval_status": "",
                "distance_to_nearest_gold": None,
                "normalized_label_min3": "",
                "normalized_label_min5": "",
                "normalized_status_min5": "",
                "match_source": "",
            }

        enriched_rows.append(
            {
                **row_dict,
                "gold_label": aligned["ground_truth_label"],
                "predicted_label": aligned["predicted_label"],
                "prediction_confidence": aligned["prediction_confidence"],
                "prediction_reason": aligned["prediction_reason"],
                "eval_status": aligned["eval_status"],
                "distance_to_nearest_gold": aligned["distance_to_nearest_gold"],
                "normalized_label_min3": aligned["normalized_label_min3"],
                "normalized_label_min5": aligned["normalized_label_min5"],
                "normalized_status_min5": aligned["normalized_status_min5"],
                "match_source": aligned["match_source"],
            }
        )

    out_df = pd.DataFrame(enriched_rows)
    args.out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(args.out_xlsx, engine="openpyxl") as writer:
        out_df.to_excel(writer, index=False, sheet_name="review")

    if not args.no_colors:
        add_conditional_colors(args.out_xlsx, status_columns=["eval_status", "normalized_status_min5"])

    print(f"Wrote: {args.out_xlsx}")
    print(f"Rows: {len(out_df)}")
    print(f"Columns added: 10")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
