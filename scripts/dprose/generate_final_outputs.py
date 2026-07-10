#!/usr/bin/env python3
"""Generate final review artifacts for the dProse full-corpus batch run.

Produces the following in ``final_outputs/``:

1. ``corpus_stats.csv``      -- one row per book with the most useful
                                scene-segmentation metrics, plus a corpus
                                aggregate row.
2. ``anomalous_books.csv``   -- books whose BORDER rate is a statistical
                                outlier (|z| >= threshold). Numeric columns are
                                filled here; the qualitative columns
                                (first_sentence_preview / likely_cause /
                                review_notes) are left for manual review.
3. ``all_sentences.csv`` / ``all_sentences.xlsx`` -- one row per sentence in
                                the Kleist-scenes column order
                                (``Sentence``, ``Phrase``, ``Text``,
                                ``is_scene_boundary``, ``scene_id``) plus
                                ``slug`` and ``model_reason``. Sentence 0 of
                                every book is forced to a scene boundary so
                                ``scene_id`` and ``is_scene_boundary`` stay
                                aligned.
4. ``per_book_xlsx/<slug>.xlsx`` -- one workbook per book with the same
                                Kleist-scenes columns + ``model_reason``.
5. ``final_report.md``       -- short plain-language corpus report.

The script reads per-book ``book_review.json`` (pre-computed scene stats) and
``predictions.jsonl`` (sentence-level labels + reasoning) under the run's
``books/`` folder, plus ``corpus_progress.json`` for the authoritative book
list and totals.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import statistics
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

# openpyxl rejects ASCII control chars (except tab/LF/CR) in cell values.
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def sanitize_xlsx_text(value: Any) -> Any:
    """Strip characters that openpyxl refuses to write into a cell."""
    if not isinstance(value, str):
        return value
    return _ILLEGAL_XLSX_RE.sub("", value)

# Reuse the per-book scene-stat logic from the review script so that books
# missing a pre-computed book_review.json (e.g. pilot-seeded books) still get
# identical metrics computed on the fly from their predictions.jsonl.
_SCRIPTS_ROOT = Path(__file__).resolve().parents[1]
if str(_SCRIPTS_ROOT) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_ROOT))
from evaluation.review_dprose_book import summarize_book  # noqa: E402

# --- Constants -------------------------------------------------------------

LONG_SCENE_MIN = 10  # a scene of >= this many sentences is an under-seg suspect
ANOMALY_Z = 2.0      # |z-score| threshold on border_rate to flag a book
DEFAULT_RUN = Path("outputs/runs/dprose_batch/dprose-full-corpus")

# Qualitative review of each anomalous book, written after reading its actual
# BORDER sentences + model reasoning in predictions.jsonl (not from thresholds).
# Keyed by slug -> (likely_cause, review_notes). Kept in code so the CSV can be
# regenerated without losing the manual analysis.
MANUAL_REVIEW: dict[str, tuple[str, str]] = {
    # --- HIGH border rate (over-segmentation suspects) ---
    "dprose_1113": (
        "Dense imagistic nature sketch ('Der Feldhase'): the model treats every "
        "small shift of weather/time/perspective as a new scene, so almost every "
        "short descriptive sentence becomes its own micro-scene.",
        "73% of scenes are 1-2 sentences and there is a 10-long BORDER run; the "
        "79 merge candidates are mostly continuations of one descriptive passage "
        "and should be merged, not treated as separate scenes.",
    ),
    "dprose_2112": (
        "Epistolary travel novel ('Kaesebiers Italienreise'): letter headers, "
        "datelines and salutations each trigger BORDER. Letter openings are real "
        "boundaries, but many mid-letter shifts are over-segmentation.",
        "Inspect the consecutive-BORDER clusters at letter boundaries (e.g. idx "
        "0-5). Letter-opening BORDERs are valid; adjacent dateline/greeting rows "
        "among the 141 merge candidates should collapse to one scene per letter.",
    ),
    "dprose_2051": (
        "Storm's frame narrative ('Eine Halligfahrt'): frequent temporal and "
        "perspective shifts (historical past vs narrator's present) plus embedded "
        "description; the model marks each stylistic shift as a new scene.",
        "Check the idx 3-6 cluster (history -> present -> first-person); many of "
        "the 102 merge candidates are shifts within one continuous meditation and "
        "are likely over-split.",
    ),
    "dprose_1913": (
        "Fairy-tale/legend with chapter headings ('Der Goldbrunnen'): a heading "
        "('I.') and the following opening sentence both fire BORDER, and each "
        "plot turn is marked as a new scene.",
        "Heading + following-sentence pairs (e.g. idx 0-2) are the clearest merge "
        "candidates among the 45; verify heading rows are not counted as separate "
        "scenes.",
    ),
    "dprose_2014": (
        "Terse, densely packed literary prose ('Ulrike') with many short "
        "sentences and rapid spatial/character micro-shifts; the model marks each "
        "setting or character shift.",
        "Check idx 7-9 (exterior -> interior -> indoor activity); the 76 merge "
        "candidates cluster around fast setting shifts inside continuous action "
        "and are probable over-segmentation.",
    ),
    "dprose_965": (
        "Keller's episodic novella narration ('Der Narr auf Manegg') with "
        "frequent explicit time markers ('Einige Zeit nach...', 'An einem "
        "schoenen Septembertage'); each temporal jump is marked BORDER.",
        "The title+first-sentence pair (idx 0-1) and other time-cue BORDERs: "
        "among the 64 merge candidates, check whether consecutive time cues are a "
        "single scene onset rather than several.",
    ),
    "dprose_1347": (
        "Dialogue-heavy frame story ('Im Wollteufel'): a round-robin where each "
        "man tells how he met his wife. Speaker turns and shifts between frame "
        "and told anecdote inflate the BORDER rate.",
        "idx 15-17 shows continuous narration being split; review the 56 merge "
        "candidates around speaker/anecdote transitions so each told story stays "
        "one scene.",
    ),
    "dprose_2013": (
        "Compressed expressionist narration ('Schuhlin', Doeblin-style): frequent "
        "'eines Tages' turning points and thematic micro-turns, each marked as a "
        "new scene. Same terse register as dprose_2007/2014.",
        "See the idx 7-9 cluster; the 57 merge candidates around dense turning "
        "points are likely over-split within one continuous account.",
    ),
    "dprose_2007": (
        "Terse, event-dense literary prose ('Die Poularde'): each new event or "
        "time marker ('im Mai', the mother's sudden death) is flagged. Same "
        "expressionist register cluster as dprose_2013/2014.",
        "See idx 5-6; among the 66 merge candidates verify whether sequential "
        "plot beats are actually one continuous scene.",
    ),
    # --- LOW border rate (under-segmentation suspects) ---
    "dprose_693": (
        "Salon detective novella from the recurring 'Dagobert Trostler' series "
        "(Grumbach household): long continuous dialogue in one setting with few "
        "physical scene changes, so BORDER rarely fires.",
        "40-sentence no-BORDER stretch at idx 116-155; scan the 187 split "
        "candidates in the long dialogue runs for an unmarked topic/time shift "
        "that may deserve a scene break.",
    ),
    "dprose_697": (
        "Same 'Dagobert Trostler' detective series ('Die feinen Zigarren'): a "
        "sustained after-dinner smoking-room conversation in one location gives "
        "long single-setting scenes.",
        "44-sentence no-BORDER stretch at idx 233-276; among the 251 split "
        "candidates check for embedded flashback/anecdote starts that should be "
        "borders.",
    ),
    "dprose_701": (
        "Same detective series ('Eine Verhaftung'): long conversational scenes "
        "with occasional embedded anecdotes. The low rate reflects genuinely long "
        "continuous talk (median scene length 5).",
        "idx 281-312 is unbroken; verify that anecdote/flashback boundaries within "
        "the dialogue are all captured among the 260 split candidates.",
    ),
    "dprose_702": (
        "Same detective series ('Empfang beim Ministerpraesidenten'): extended "
        "witty mock-trial dialogue in a single setting produces long scenes.",
        "40-sentence stretch at idx 77-116; among the 304 split candidates check "
        "topic shifts in the long dialogue for missed borders.",
    ),
    "dprose_1075": (
        "Frame story with one long embedded first-person tale ('Wie der Teufel "
        "den Professor holte'): sustained storytelling to a group keeps a single "
        "scene running, so the BORDER rate is low.",
        "56-sentence unbroken stretch at idx 195-250 (inside the told story); "
        "check the 272 split candidates for frame<->tale shifts that may need "
        "borders.",
    ),
    "dprose_2006": (
        "Continuous character-focused narrative built around one relationship "
        "('Die Laus' / the woman and the beggar); few discrete scene changes give "
        "long scenes.",
        "46-sentence stretch at idx 10-55; scan the 304 split candidates for "
        "unmarked time jumps (letter received 'gestern', the beggar's arrival) "
        "that could be borders.",
    ),
    "dprose_661": (
        "Long single-protagonist novella ('Der Wurzgartner', a judge's holiday in "
        "Tirol) told as continuous narration with sustained descriptive and "
        "reflective passages; the lowest BORDER rate in the corpus.",
        "56-sentence no-BORDER stretch at idx 120-175; with 408 split candidates "
        "(most in the corpus) this is the prime book for a manual pass to find "
        "missed scene breaks in the long descriptive runs.",
    ),
}


# --- IO helpers ------------------------------------------------------------

def load_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            if line.strip():
                rows.append(json.loads(line))
    return rows


def load_json(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def slug_from_source(source_file: str) -> str:
    return source_file[:-4] if source_file.endswith(".csv") else source_file


def extract_reason(raw_model_response: str | None) -> str:
    """Pull the free-text ``reason`` out of the raw model JSON string."""
    if not raw_model_response:
        return ""
    try:
        obj = json.loads(raw_model_response)
    except (json.JSONDecodeError, TypeError):
        return ""
    if isinstance(obj, dict):
        return str(obj.get("reason", "") or "")
    return ""


# --- Per-sentence scene geometry ------------------------------------------

def scene_length_by_index(border_flags: list[bool]) -> list[int]:
    """For each sentence, the length (in sentences) of the scene it belongs to.

    A scene starts at a BORDER sentence and runs until the sentence before the
    next BORDER. If the first sentence is not a BORDER it still opens scene 0.
    """
    n = len(border_flags)
    if n == 0:
        return []
    starts = [i for i, is_border in enumerate(border_flags) if is_border]
    if not starts or starts[0] != 0:
        starts = [0] + starts
    starts = sorted(set(starts))
    lengths = [0] * n
    for si, start in enumerate(starts):
        end = starts[si + 1] if si + 1 < len(starts) else n
        for i in range(start, end):
            lengths[i] = end - start
    return lengths


def force_opening_border(border_flags: list[bool]) -> list[bool]:
    """Ensure sentence 0 is a scene boundary.

    A book's first sentence always opens scene 1; ``NOBORDER`` on index 0 is
    meaningless (continuation of nothing). Forcing True keeps
    ``is_scene_boundary`` and ``scene_id`` aligned for every book.
    """
    if not border_flags:
        return border_flags
    forced = list(border_flags)
    forced[0] = True
    return forced


def cumulative_scene_ids(border_flags: list[bool]) -> list[int]:
    """Cumulative 1-based scene index per sentence.

    Requires ``border_flags[0] is True`` (see ``force_opening_border``).
    Every subsequent BORDER=True increments the counter, so
    ``max(scene_id) == sum(border_flags)``.
    """
    scene_ids: list[int] = []
    current = 0
    for is_border in border_flags:
        if is_border:
            current += 1
        scene_ids.append(current)
    return scene_ids


def per_sentence_flags(rows_sorted: list[dict[str, Any]]) -> list[dict[str, int]]:
    """Compute merge/split/review flags + scene_id; force opening border."""
    border_flags = force_opening_border(
        [bool(r.get("prediction_bool")) for r in rows_sorted]
    )
    scene_lens = scene_length_by_index(border_flags)
    scene_ids = cumulative_scene_ids(border_flags)
    n = len(rows_sorted)
    out: list[dict[str, int]] = []
    for i in range(n):
        is_border = border_flags[i]
        prev_border = border_flags[i - 1] if i > 0 else False
        next_border = border_flags[i + 1] if i + 1 < n else False
        # Over-segmentation suspect: a BORDER adjacent to another BORDER.
        merge = int(is_border and (prev_border or next_border))
        # Under-segmentation suspect: a NOBORDER buried inside a long scene.
        split = int((not is_border) and scene_lens[i] >= LONG_SCENE_MIN)
        out.append({
            "is_scene_boundary": int(is_border),
            "merge_candidate": merge,
            "split_candidate": split,
            "review_flag": int(bool(merge or split)),
            "scene_id": scene_ids[i],
        })
    return out


# --- Book discovery --------------------------------------------------------

def book_dirs(run_dir: Path) -> dict[str, Path]:
    progress = load_json(run_dir / "corpus_progress.json")
    slugs = sorted(progress.get("books", {}).keys())
    dirs: dict[str, Path] = {}
    for slug in slugs:
        d = run_dir / "books" / slug
        if d.is_dir():
            dirs[slug] = d
    return dirs


# --- Artifact 1: corpus_stats.csv -----------------------------------------

CORPUS_STAT_FIELDS = [
    "slug",
    "sentence_count",
    "parse_ok_rate",
    "border_rate",
    "scene_length_median",
    "scene_length_mean",
    "scene_length_min",
    "scene_length_max",
    "short_scene_rate",
    "consecutive_border_pairs",
    "max_consecutive_border_run",
    "gaps_ge_10",
    "estimated_cost_usd",
]


def build_corpus_stats(reviews: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for slug in sorted(reviews):
        rv = reviews[slug]
        rows.append({
            "slug": slug,
            "sentence_count": rv.get("sentence_count", 0),
            "parse_ok_rate": round(float(rv.get("parse_ok_rate", 0.0)), 4),
            "border_rate": round(float(rv.get("border_rate", 0.0)), 4),
            "scene_length_median": rv.get("scene_length_median", 0),
            "scene_length_mean": round(float(rv.get("scene_length_mean", 0.0)), 2),
            "scene_length_min": rv.get("scene_length_min", 0),
            "scene_length_max": rv.get("scene_length_max", 0),
            "short_scene_rate": round(float(rv.get("short_scene_rate", 0.0)), 4),
            "consecutive_border_pairs": rv.get("consecutive_border_pairs", 0),
            "max_consecutive_border_run": rv.get("max_consecutive_border_run", 0),
            "gaps_ge_10": rv.get("gaps_ge_10", 0),
            "estimated_cost_usd": round(float(rv.get("estimated_cost_usd", 0.0)), 4),
        })
    return rows


def aggregate_row(rows: list[dict[str, Any]]) -> dict[str, Any]:
    def col(name: str) -> list[float]:
        return [float(r[name]) for r in rows]

    total_sent = sum(int(r["sentence_count"]) for r in rows)
    total_cost = sum(float(r["estimated_cost_usd"]) for r in rows)
    return {
        "slug": f"__CORPUS_AGGREGATE__ (n_books={len(rows)})",
        "sentence_count": total_sent,
        "parse_ok_rate": round(statistics.mean(col("parse_ok_rate")), 4),
        "border_rate": round(statistics.mean(col("border_rate")), 4),
        "scene_length_median": round(statistics.median(col("scene_length_median")), 2),
        "scene_length_mean": round(statistics.mean(col("scene_length_mean")), 2),
        "scene_length_min": min(int(r["scene_length_min"]) for r in rows),
        "scene_length_max": max(int(r["scene_length_max"]) for r in rows),
        "short_scene_rate": round(statistics.mean(col("short_scene_rate")), 4),
        "consecutive_border_pairs": sum(int(r["consecutive_border_pairs"]) for r in rows),
        "max_consecutive_border_run": max(int(r["max_consecutive_border_run"]) for r in rows),
        "gaps_ge_10": sum(int(r["gaps_ge_10"]) for r in rows),
        "estimated_cost_usd": round(total_cost, 4),
    }


def write_corpus_stats(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CORPUS_STAT_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
        writer.writerow(aggregate_row(rows))


# --- Artifact 2: anomalous_books.csv --------------------------------------

ANOMALY_FIELDS = [
    "slug",
    "sentence_count",
    "border_rate",
    "border_rate_zscore",
    "anomaly_type",
    "scene_length_median",
    "short_scene_rate",
    "max_consecutive_border_run",
    "consecutive_border_pairs",
    "gaps_ge_10",
    "merge_candidate_count",
    "split_candidate_count",
    # Qualitative columns filled during manual review:
    "first_sentence_preview",
    "likely_cause",
    "review_notes",
]


def build_anomalous(
    reviews: dict[str, dict[str, Any]],
    candidate_counts: dict[str, dict[str, Any]],
    z_threshold: float,
) -> list[dict[str, Any]]:
    rates = {s: float(rv.get("border_rate", 0.0)) for s, rv in reviews.items()}
    mean = statistics.mean(rates.values())
    stdev = statistics.pstdev(rates.values()) or 1.0

    anomalies: list[dict[str, Any]] = []
    for slug, rate in rates.items():
        z = (rate - mean) / stdev
        if abs(z) < z_threshold:
            continue
        rv = reviews[slug]
        counts = candidate_counts.get(slug, {})
        anomalies.append({
            "slug": slug,
            "sentence_count": rv.get("sentence_count", 0),
            "border_rate": round(rate, 4),
            "border_rate_zscore": round(z, 2),
            "anomaly_type": "HIGH" if z > 0 else "LOW",
            "scene_length_median": rv.get("scene_length_median", 0),
            "short_scene_rate": round(float(rv.get("short_scene_rate", 0.0)), 4),
            "max_consecutive_border_run": rv.get("max_consecutive_border_run", 0),
            "consecutive_border_pairs": rv.get("consecutive_border_pairs", 0),
            "gaps_ge_10": rv.get("gaps_ge_10", 0),
            "merge_candidate_count": counts.get("merge_candidate", 0),
            "split_candidate_count": counts.get("split_candidate", 0),
            "first_sentence_preview": counts.get("first_opening")
            or rv.get("first_sentence_preview", ""),
            "likely_cause": MANUAL_REVIEW.get(slug, ("", ""))[0],
            "review_notes": MANUAL_REVIEW.get(slug, ("", ""))[1],
        })
    anomalies.sort(key=lambda r: r["border_rate_zscore"], reverse=True)
    return anomalies


def write_anomalous(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=ANOMALY_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


# --- Artifact 3: all_sentences.csv / .xlsx ---------------------------------
# Column order matches Scenes_example output.xlsx, plus slug (corpus files)
# and model_reason. dProse has no phrase split, so Phrase == 0-based index
# and Sentence == index + 1 (1-based), matching the example's numbering roles.

SENTENCE_FIELDS = [
    "slug",
    "Sentence",
    "Phrase",
    "Text",
    "is_scene_boundary",
    "scene_id",
    "model_reason",
]

# Per-book xlsx: unnamed leading index column like the Kleist example, then
# the same named columns (no slug — one book per file) + model_reason.
PER_BOOK_XLSX_HEADERS = [
    None,
    "Sentence",
    "Phrase",
    "Text",
    "is_scene_boundary",
    "scene_id",
    "model_reason",
]

CORPUS_XLSX_HEADERS = [
    "slug",
    "Sentence",
    "Phrase",
    "Text",
    "is_scene_boundary",
    "scene_id",
    "model_reason",
]


def process_book_sentences(
    slug: str, rows: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Return (sentence rows for export, per-book aggregates).

    Sentence 0 is forced to a scene boundary so ``is_scene_boundary`` and
    ``scene_id`` stay aligned. Aggregates include merge/split candidate
    totals (for anomalous_books.csv) and the true opening-sentence preview.
    """
    rows_sorted = sorted(rows, key=lambda r: int(r.get("sentence_index", 0)))
    flags = per_sentence_flags(rows_sorted)

    out_rows: list[dict[str, Any]] = []
    merge_total = 0
    split_total = 0
    forced_opening = 0
    for row, flag in zip(rows_sorted, flags):
        merge_total += flag["merge_candidate"]
        split_total += flag["split_candidate"]
        idx = int(row.get("sentence_index", 0))
        model_border = bool(row.get("prediction_bool"))
        if idx == 0 and not model_border:
            forced_opening = 1
        out_rows.append({
            "slug": slug,
            "index": idx,
            "Sentence": idx + 1,
            "Phrase": idx,
            "Text": row.get("sentence_text_full", ""),
            "is_scene_boundary": bool(flag["is_scene_boundary"]),
            "scene_id": flag["scene_id"],
            "model_reason": extract_reason(row.get("raw_model_response")),
            # Internal fields used by report / anomaly aggregates only:
            "review_flag": flag["review_flag"],
            "manual_fix": row.get("manual_fix", ""),
        })
    opening = rows_sorted[0].get("sentence_text_full", "")[:120] if rows_sorted else ""
    return out_rows, {
        "merge_candidate": merge_total,
        "split_candidate": split_total,
        "first_opening": opening,
        "forced_opening_border": forced_opening,
    }


def write_all_sentences_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SENTENCE_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({
                **row,
                # CSV cannot store real booleans; write TRUE/FALSE so Excel
                # imports them as Wahr/Falsch under a German locale.
                "is_scene_boundary": "TRUE" if row["is_scene_boundary"] else "FALSE",
            })


def _style_text_column(ws: Any, col_letter: str) -> None:
    for cell in ws[col_letter][1:]:
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")


def write_all_sentences_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    """Corpus-wide workbook with the same columns as all_sentences.csv."""
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("scenes")
    # write_only Workbooks start with no usable default sheet; create_sheet
    # is the supported entry point.
    ws.append(CORPUS_XLSX_HEADERS)
    for row in rows:
        ws.append([
            row["slug"],
            int(row["Sentence"]),
            int(row["Phrase"]),
            sanitize_xlsx_text(row["Text"]),
            bool(row["is_scene_boundary"]),
            int(row["scene_id"]),
            sanitize_xlsx_text(row["model_reason"]),
        ])
    wb.save(path)


def write_per_book_xlsx(out_dir: Path, all_sentence_rows: list[dict[str, Any]]) -> int:
    """Write one .xlsx per book in the Kleist-scenes example layout + model_reason.

    Columns: unnamed 0-based index, Sentence (1-based), Phrase (0-based unit
    index; equals Phrase==index because dProse has no phrase split), Text,
    is_scene_boundary (Excel bool -> Wahr/Falsch), scene_id, model_reason.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    by_slug: dict[str, list[dict[str, Any]]] = {}
    for row in all_sentence_rows:
        by_slug.setdefault(row["slug"], []).append(row)

    for slug, rows in by_slug.items():
        rows_sorted = sorted(rows, key=lambda r: int(r["index"]))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "scenes"
        ws.append(PER_BOOK_XLSX_HEADERS)
        for row in rows_sorted:
            ws.append([
                int(row["index"]),
                int(row["Sentence"]),
                int(row["Phrase"]),
                sanitize_xlsx_text(row["Text"]),
                bool(row["is_scene_boundary"]),
                int(row["scene_id"]),
                sanitize_xlsx_text(row["model_reason"]),
            ])
        ws.freeze_panes = "A2"
        widths = {"A": 8, "B": 10, "C": 10, "D": 100, "E": 18, "F": 10, "G": 60}
        for letter, width in widths.items():
            ws.column_dimensions[letter].width = width
        _style_text_column(ws, get_column_letter(4))
        _style_text_column(ws, get_column_letter(7))
        wb.save(out_dir / f"{slug}.xlsx")

    return len(by_slug)


# --- Artifact 4: final_report.md ------------------------------------------

def write_report(
    path: Path,
    *,
    progress: dict[str, Any],
    corpus_rows: list[dict[str, Any]],
    anomalies: list[dict[str, Any]],
    total_sentences_seen: int,
    total_review_flags: int,
    total_manual_fixes: int,
) -> None:
    totals = progress.get("totals", {})
    n_books = totals.get("books_complete", len(corpus_rows))
    n_sent = totals.get("sentences_complete", total_sentences_seen)
    cost = totals.get("cost_usd", 0.0)

    border_rates = sorted(float(r["border_rate"]) for r in corpus_rows)
    median_border = statistics.median(border_rates)
    min_border, max_border = border_rates[0], border_rates[-1]
    parse_rates = [float(r["parse_ok_rate"]) for r in corpus_rows]
    mean_parse = statistics.mean(parse_rates)

    scene_medians = [float(r["scene_length_median"]) for r in corpus_rows]
    corpus_scene_median = statistics.median(scene_medians)
    short_rates = [float(r["short_scene_rate"]) for r in corpus_rows]
    mean_short = statistics.mean(short_rates)

    n_high = sum(1 for a in anomalies if a["anomaly_type"] == "HIGH")
    n_low = sum(1 for a in anomalies if a["anomaly_type"] == "LOW")

    lines = [
        "# dProse Full-Corpus Scene Segmentation — Final Report",
        "",
        "## What we did",
        "",
        "We ran automatic scene-boundary labelling over the whole dProse prose "
        "corpus with Gemini 2.5 Pro (batch API, prompt family B, 12 sentences of "
        "context each side). Every sentence gets one of two labels: `BORDER` (a new "
        "scene starts here) or `NOBORDER` (same scene continues). Because dProse "
        "has no gold scene labels, we could not measure accuracy directly. Instead "
        "we ran the corpus in cost-capped waves and, after each wave, checked the "
        "health and plausibility of the output: parse success, BORDER rate, scene "
        "lengths, and runs of consecutive BORDERs. Books with unusual numbers were "
        "read by hand around the flagged sentences. Sentences the model failed to "
        "label (thinking overflow, transient API errors, safety blocks) were fixed "
        "in three tiers: batch retry, synchronous retry, and finally a "
        "neighbour-agreement patch for the last few stubborn cases. The full log "
        "lives in `docs/corpora/DPROSE_CORPUS_SPOT_CHECKS.md`.",
        "",
        "## Corpus at a glance",
        "",
        f"- **{n_books} books**, **{n_sent:,} sentences** labelled end to end.",
        f"- Total estimated cost: **~${cost:,.0f} USD**.",
        f"- Parse success after remediation: **{mean_parse:.1%}** mean per book "
        f"(all {total_manual_fixes} unresolved sentences patched by neighbour "
        "agreement).",
        f"- Median BORDER rate across books: **{median_border:.1%}** "
        f"(range {min_border:.1%} to {max_border:.1%}).",
        f"- Statistical outliers on BORDER rate (|z| >= {ANOMALY_Z:.0f}): "
        f"**{len(anomalies)} books** ({n_high} high, {n_low} low).",
        "",
        "## Scene segmentation patterns",
        "",
        f"- Typical scene length is short: corpus-median of the per-book median "
        f"scene length is **{corpus_scene_median:.0f} sentences**.",
        f"- On average **{mean_short:.0%}** of inferred scenes are only 1-2 "
        "sentences long — the main over-segmentation signal to watch.",
        f"- Across the corpus, **{total_review_flags:,} sentences** "
        f"({total_review_flags / max(1, total_sentences_seen):.1%}) are flagged "
        "for manual review (either merge or split candidates).",
        "- Runs of consecutive BORDERs (several scene breaks in a row) are the "
        "clearest sign the model over-split a passage, usually around dialogue.",
        "",
        "## Nature of the corpus",
        "",
        "dProse is heterogeneous German prose: short lyrical pieces sit next to "
        "long novels, and registers range from fairy-tale narration to dense "
        "epistolary and dialogue-heavy texts. This shows up directly in the wide "
        "spread of BORDER rates: there is no single \"correct\" rate for the corpus, "
        "so a book's rate is only meaningful relative to its own genre and length. "
        "High-rate books tend to be fragmentary or dialogue-driven (many short "
        "scenes); low-rate books tend to be single continuous passages (letters, "
        "monologue, sustained description). The per-book metrics in "
        "`corpus_stats.csv` and the outlier notes in `anomalous_books.csv` are the "
        "right lens for judging any individual book.",
        "",
        "## Known data-quality notes",
        "",
        f"- **{total_manual_fixes} sentences** were labelled by neighbour "
        "agreement (not by the model) after retries failed; they live in the "
        "per-book `predictions.jsonl` under `manual_fix=neighbor_agreement`.",
        "- A few sentences were originally blocked by the model's safety filter "
        "(`PROHIBITED_CONTENT`); these are among the patched rows.",
        "- Sentence 0 of every book is forced to `is_scene_boundary=True` so "
        "`scene_id` and the boundary flag stay aligned (a book's first sentence "
        "always opens scene 1).",
        "- The older `corpus_summary.json` and `predictions_full.jsonl` in the run "
        "root are **stale pilot artifacts** (3 books only) and should be ignored; "
        "the authoritative state is `corpus_progress.json` plus the per-book "
        "files under `books/`.",
        "",
        "## Files in this folder",
        "",
        "- `corpus_stats.csv` — per-book metrics + a corpus aggregate row.",
        "- `anomalous_books.csv` — BORDER-rate outliers with review guidance.",
        "- `all_sentences.csv` / `all_sentences.xlsx` — every sentence in the "
        "Kleist-scenes column order (`Sentence`, `Phrase`, `Text`, "
        "`is_scene_boundary`, `scene_id`) plus `slug` and `model_reason`.",
        "- `per_book_xlsx/<slug>.xlsx` — one workbook per book with the same "
        "Kleist-scenes columns + `model_reason` (`is_scene_boundary` as "
        "True/False → Wahr/Falsch in German Excel).",
        "- `final_report.md` — this file.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


# --- Main ------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--run_dir", type=Path, default=DEFAULT_RUN)
    parser.add_argument("--z_threshold", type=float, default=ANOMALY_Z)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    run_dir: Path = args.run_dir
    out_dir = run_dir / "final_outputs"
    out_dir.mkdir(parents=True, exist_ok=True)

    progress = load_json(run_dir / "corpus_progress.json")
    dirs = book_dirs(run_dir)
    print(f"Found {len(dirs)} book folders under {run_dir / 'books'}")

    reviews: dict[str, dict[str, Any]] = {}
    candidate_counts: dict[str, dict[str, Any]] = {}
    all_sentence_rows: list[dict[str, Any]] = []
    total_review_flags = 0
    total_manual_fixes = 0
    total_forced_opening = 0

    for slug in sorted(dirs):
        book_dir = dirs[slug]
        review_path = book_dir / "book_review.json"
        pred_path = book_dir / "predictions.jsonl"
        if not pred_path.exists():
            print(f"  WARN: no predictions.jsonl for {slug}, skipping")
            continue

        pred_rows = load_jsonl(pred_path)
        if review_path.exists():
            reviews[slug] = load_json(review_path)
        else:
            # Pilot-seeded books lack a book_review.json; compute it on the fly.
            print(f"  INFO: computing stats for {slug} (no book_review.json)")
            reviews[slug] = summarize_book(pred_rows, source_file=f"{slug}.csv")
        sent_rows, counts = process_book_sentences(slug, pred_rows)
        candidate_counts[slug] = counts
        all_sentence_rows.extend(sent_rows)
        total_review_flags += sum(r["review_flag"] for r in sent_rows)
        total_manual_fixes += sum(1 for r in sent_rows if r["manual_fix"])
        total_forced_opening += int(counts.get("forced_opening_border", 0))

    all_sentence_rows.sort(key=lambda r: (r["slug"], r["index"]))

    corpus_rows = build_corpus_stats(reviews)
    anomalies = build_anomalous(reviews, candidate_counts, args.z_threshold)

    write_corpus_stats(out_dir / "corpus_stats.csv", corpus_rows)
    write_anomalous(out_dir / "anomalous_books.csv", anomalies)
    write_all_sentences_csv(out_dir / "all_sentences.csv", all_sentence_rows)
    write_all_sentences_xlsx(out_dir / "all_sentences.xlsx", all_sentence_rows)
    n_xlsx = write_per_book_xlsx(out_dir / "per_book_xlsx", all_sentence_rows)
    write_report(
        out_dir / "final_report.md",
        progress=progress,
        corpus_rows=corpus_rows,
        anomalies=anomalies,
        total_sentences_seen=len(all_sentence_rows),
        total_review_flags=total_review_flags,
        total_manual_fixes=total_manual_fixes,
    )

    print("\n=== Done ===")
    print(f"  corpus_stats.csv     : {len(corpus_rows)} book rows (+1 aggregate)")
    print(f"  anomalous_books.csv  : {len(anomalies)} outlier books")
    print(f"  all_sentences.csv    : {len(all_sentence_rows):,} sentence rows")
    print(f"  all_sentences.xlsx   : {len(all_sentence_rows):,} sentence rows")
    print(f"  per_book_xlsx/       : {n_xlsx} workbooks")
    print(f"  review flags         : {total_review_flags:,}")
    print(f"  manual_fix rows      : {total_manual_fixes}")
    print(f"  forced opening border: {total_forced_opening} books")
    missing_review = [a["slug"] for a in anomalies if a["slug"] not in MANUAL_REVIEW]
    if missing_review:
        print(
            "\nWARNING: these flagged books have no MANUAL_REVIEW entry "
            f"(qualitative columns left blank): {missing_review}"
        )

    print("\nAnomalous slugs:")
    for a in anomalies:
        print(
            f"  {a['slug']:<16} {a['anomaly_type']:<4} "
            f"border={a['border_rate']:.3f} z={a['border_rate_zscore']:+.2f} "
            f"merge={a['merge_candidate_count']} split={a['split_candidate_count']}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
